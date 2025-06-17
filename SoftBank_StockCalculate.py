import pandas as pd
from datetime import datetime, timedelta
from sqlalchemy import create_engine
import sys
import logging
import os  # Add os module to handle paths
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import calendar

# Configure logging
log_filename = f"logfile_{datetime.now().strftime('%Y-%m-%d')}.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename, encoding='utf-8'),
        logging.StreamHandler()
    ]
)


# Connect to MSSQL
def connect_to_db(server, database):
    try:
        conn_str = f"mssql+pyodbc://@{server}/{database}?driver=ODBC+Driver+17+for+SQL+Server&trusted_connection=yes"
        engine = create_engine(conn_str)
        return engine
    except Exception as e:
        logging.error(f"資料庫連接失敗: {e}")
        raise


# Fetch data
def fetch_data():
    try:
        engine = connect_to_db('jpdejitdev01', 'ITQAS2')
        queries = {
            'factory_data': "SELECT Part_No, eta_FLTC, Qty, Status FROM SoftBank_Data_FactoryShipment",
            'order_data': "SELECT Product_Name, COALESCE(Actual_shipment_Date, Estimated_Shipment_Date) AS Shipment_Date, Quantity, Quotation_status FROM SoftBank_Data_Orderinfo",
            'product_data': "SELECT Delta_PartNO AS Part_No, [Month-End_SAP_Inventory], Model FROM SoftBank_Data_Productinfo"
        }
        return {name: pd.read_sql(query, engine) for name, query in queries.items()}
    except Exception as e:
        logging.error(f"Error fetching data: {e}")
        raise

def calculate_inventory(factory_data, order_data, product_data):
    try:
        # print("=== 開始庫存計算調試 ===")
        
        # Clean data
        factory_data['eta_FLTC'] = pd.to_datetime(factory_data['eta_FLTC']).dt.date
        order_data['Shipment_Date'] = pd.to_datetime(order_data['Shipment_Date']).dt.date

        # Perform inventory calculation(half year count)
        start_date = datetime.today().replace(day=1).date()
        end_date = start_date + timedelta(days=180)
        date_range = pd.date_range(start=start_date, end=end_date).date

        # print(f"日期範圍: {start_date} 到 {end_date}")

        # 過濾訂單數據
        order_data = order_data[
            (order_data['Shipment_Date'] >= start_date) & 
            (~order_data['Quotation_status'].isin(['quotation', 'cancel','confirming','double cancel']))
        ]

        # 獲取所有產品列表
        products = product_data['Part_No'].unique()
        # print(f"總產品數量: {len(products)}")
        
        # 創建庫存DataFrame
        inventory = pd.DataFrame(index=date_range, columns=products, data=0.0)
        # print(f"初始庫存DataFrame形狀: {inventory.shape}")

        # 設定初始庫存
        # print("\n=== 設定初始庫存 ===")
        
        # 清理product_data，移除重複和空值
        clean_product_data = product_data.dropna(subset=['Part_No', 'Month-End_SAP_Inventory'])
        clean_product_data = clean_product_data.drop_duplicates(subset=['Part_No'], keep='first')
        
        # print(f"清理後的product_data數量: {len(clean_product_data)}")
        
        # 逐個設定初始庫存
        for product in products:
            matching_rows = clean_product_data[clean_product_data['Part_No'] == product]
            
            if len(matching_rows) > 0:
                initial_qty = matching_rows['Month-End_SAP_Inventory'].iloc[0]
                # 確保是數值類型
                try:
                    initial_qty = float(initial_qty) if pd.notna(initial_qty) else 0.0
                    inventory.loc[start_date, product] = initial_qty
                except (ValueError, TypeError):
                    inventory.loc[start_date, product] = 0.0
            else:
                inventory.loc[start_date, product] = 0.0
        
        # 驗證第一天庫存設定（保留關鍵驗證）
        first_day_stock = inventory.loc[start_date]
        non_zero_count = (first_day_stock > 0).sum()
        print(f"第一天有庫存的產品數量: {non_zero_count}")
        
        # 處理工廠數據
        # print("\n=== 處理工廠進貨數據 ===")
        valid_factory_data = factory_data[factory_data['Part_No'].isin(products)]
        # print(f"有效工廠數據筆數: {len(valid_factory_data)}")
        
        if not valid_factory_data.empty:
            factory_agg = valid_factory_data.groupby(['eta_FLTC', 'Part_No'])['Qty'].sum().unstack(fill_value=0)
            factory_agg = factory_agg.reindex(columns=products, index=date_range, fill_value=0)
        else:
            factory_agg = pd.DataFrame(0, index=date_range, columns=products)
        
        # 處理訂單數據
        # print("\n=== 處理訂單出貨數據 ===")
        # print(f"有效訂單數據筆數: {len(order_data)}")
        
        if not order_data.empty:
            order_agg = order_data.groupby(['Shipment_Date', 'Product_Name'])['Quantity'].sum().unstack(fill_value=0)
            order_agg = order_agg.reindex(columns=products, index=date_range, fill_value=0)
        else:
            order_agg = pd.DataFrame(0, index=date_range, columns=products)

        # 庫存計算邏輯
        # print("\n=== 執行庫存計算 ===")
        
        # 逐日計算庫存
        for i, current_date in enumerate(date_range):
            if i == 0:
                # 第一天：初始庫存 + 當天進貨 - 當天出貨
                daily_in = factory_agg.loc[current_date]
                daily_out = order_agg.loc[current_date]
                inventory.loc[current_date] = inventory.loc[current_date] + daily_in - daily_out
            else:
                # 其他天：前一天庫存 + 當天進貨 - 當天出貨
                prev_date = date_range[i-1]
                daily_in = factory_agg.loc[current_date]
                daily_out = order_agg.loc[current_date]
                inventory.loc[current_date] = inventory.loc[prev_date] + daily_in - daily_out
        
        # print("=== 庫存計算完成 ===\n")
        
        return inventory
        
    except Exception as e:
        logging.error(f"Error during inventory calculation: {e}")
        raise

# Export results to Excel
def export_to_excel(inventory, product_data):
    try:
        """
        Export calculated inventory to Excel, merging equivalent Part_No before exporting
        """
        excluded_parts = ['DEJ-OR-FRT-01399','出力定格電圧調整費(54.6V→52.8V)',
                          '3073247220','(BB)TBM48050E2-1M22','(BA)TBM48050E2-1M22','(C)TBM48050E2-1S22']
        print(f"開始處理庫存數據，將排除以下產品: {excluded_parts}")
        
        # 第一步：排除指定產品
        for part_no in excluded_parts:
            if part_no in inventory.columns:
                inventory.drop(columns=[part_no], inplace=True)
                print(f"  ✓ 已排除產品: {part_no}")
            else:
                print(f"  - 未找到要排除的產品: {part_no}")
        # 1對1合併
        part_no_mapping = {
            '3798D000000278-S(free)': '3798D000000278-S',
            '3798D000000225-S(free)': '3798D000000225-S',
            '3798D000000228-S(free)': '3798D000000228-S',
            'ESR-48/56L J-S(free)': 'ESR-48/56L J-S',
            'ESBC200-CEA04(supplied materials)': 'ESBC200-CEA04',
            'ESR-48/56C F-A(free)': 'ESR-48/56C F-A',
            'ESAA75-CEA03(supplied materials)': 'ESAA75-CEA03',
            'ESOF040-EAA01(supplied materials)': 'ESOF040-EAA01',
            '3798D000000763-S(supplied materials)': '3798D000000763-S',
            '3798D000000762-S(supplied materials)': '3798D000000762-S',
            '3798D000000761-S(supplied materials)': '3798D000000761-S',
            '3798D000000760-S(supplied materials)': '3798D000000760-S',
            '3798D000000764-S(supplied materials)': '3798D000000764-S',
            '3798C000000642-S(supplied materials)': '3798C000000642-S',
            '3798D000000805-S(supplied materials)': '3798D000000805-S',
            '3798D000000806-S(supplied materials)': '3798D000000806-S',
            '3798Z00099AT-S(supplied materials)': '3798Z00099AT-S',
            '3798C000000620-S(supplied materials)': '3798C000000620-S',
            '3798C000000621-S(supplied materials)': '3798C000000621-S',
            '3798D000000315-S(free)':'3798D000000315-S',
            'ESAA75-CEA02(supplied materials)':'ESAA75-CEA02',
            '3377144600-S(free)':'3377144600-S',
            '3474179500(free)':'3474179500'
        }

        # 1對1合併
        for free_part_no, main_part_no in part_no_mapping.items():
            if free_part_no in inventory.columns:
                if main_part_no in inventory.columns:
                    inventory[main_part_no] += inventory[free_part_no]
                else:
                    inventory.rename(columns={free_part_no: main_part_no}, inplace=True)
                inventory.drop(columns=[free_part_no], inplace=True, errors='ignore')

        # 多對1合併
        multi_mapping = {
            '3798C000000622-S': [ 
                '3798C000000622-S(free)',  
                '3798C000000622-S(supplied materials)(free)',  
                '3798C000000622-S(supplied materials)'  
            ],
            
            '3799906300-S': [
                '3799906300-S(free)',
                '3799906300-S(supplied materials)(free)'
            ],
            
            '3799906200-S': [
                '3799906200-S(free)',
                '3799906200-S(supplied materials)(free)'
            ],
            
            'ESBC200-CEA01': [
                'ESBC200-CEA01(supplied materials)',
                'ESBC200-CEA02(ESBC200-CEA01rework)',
                'ESBC200-CEA03(ESBC200-CEA01rework)',
                'ESBC200-CEA04(ESBC200-CEA01rework)',
                'ESBC200-CEA02(ESBC200-CEA01rework supplied materials)',
                'ESBC200-CEA03(ESBC200-CEA01rework supplied materials)',
                'ESBC200-CEA04(ESBC200-CEA01rework supplied materials)'
            ],
            
            'ESBC200-CEA02': [
                'ESBC200-CEA02(supplied materials)',
                'ESBC200-CEA03(ESBC200-CEA02rework)',
                'ESBC200-CEA04(ESBC200-CEA02rework)',
                'ESBC200-CEA03(ESBC200-CEA02rework supplied materials)',
                'ESBC200-CEA04(ESBC200-CEA02rework supplied materials)'
            ],
            
            'ESBC200-CEA03': [
                'ESBC200-CEA03(supplied materials)',
                'ESBC200-CEA04(ESBC200-CEA03rework)',
                'ESBC200-CEA04(ESBC200-CEA03rework supplied materials)'
            ],
            
            'ESBC200-CEA05': [
                'ESBC200-CEA01(ESBC200-CEA05rework)',
                'ESBC200-CEA02(ESBC200-CEA05rework)',
                'ESBC200-CEA03(ESBC200-CEA05rework)',
                'ESBC200-CEA04(ESBC200-CEA05rework)',
                'ESBC200-CEA01(ESBC200-CEA05rework supplied materials)',
                'ESBC200-CEA02(ESBC200-CEA05rework supplied materials)',
                'ESBC200-CEA03(ESBC200-CEA05rework supplied materials)',
                'ESBC200-CEA04(ESBC200-CEA05rework supplied materials)'
            ],
            'ESAA75-CEA01': [
                'ESAA75-CEA01(supplied materials)',
                'ESAA75-CEA02(ESAA75-CEA01rework supplied materials)',
                'ESAA75-CEA02(ESAA75-CEA01rework)'
            ],
                
                'ESAA75-CEA04': [
                'ESAA75-CEA04(supplied materials)',
                'ESAA75-CEA03(ESAA75-CEA04rework)',
                'ESAA75-CEA03(ESAA75-CEA04rework supplied materials)'
            ],
                'ESAA75-CEA05': [
                'ESAA75-CEA05(supplied materials)',
                'ESAA75-CEA03(ESAA75-CEA05rework)',
                'ESAA75-CEA04(ESAA75-CEA05rework)',
                'ESAA75-CEA03(ESAA75-CEA05rework supplied materials)',
                'ESAA75-CEA04(ESAA75-CEA05rework supplied materials)'
            ]
        }

        # 多對1合併
        for main_part_no, part_list in multi_mapping.items():
            for part in part_list:
                if part in inventory.columns:
                    if main_part_no in inventory.columns:
                        inventory[main_part_no] += inventory[part]
                    else:
                        inventory[main_part_no] = inventory[part]
                    inventory.drop(columns=[part], inplace=True, errors='ignore')

        # 轉置 DataFrame
        inventory_transposed = inventory.T

        # 獲取 Part_No 與 Model 對應資訊
        part_no_model_mapping = product_data.set_index('Part_No')['Model'].to_dict()

        # 在 Part_No 前面新增 Model 欄位
        inventory_transposed = inventory.T
        inventory_transposed.insert(0, 'Model', inventory_transposed.index.map(part_no_model_mapping))

        # 生成檔名
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        file_name = f"Daily_Inventory_Simulate_{timestamp}.xlsx"

        # 設定網路路徑
        save_path = r"\\jpdejstcfs01\STC_share\JP IT\STC SBK 仕分けリスト\IT system\Report"
        full_path = os.path.join(save_path, file_name)

        # 儲存為 Excel
        inventory_transposed.to_excel(full_path, index=True)

        # 讀取 Excel 進行格式化
        wb = load_workbook(full_path)
        ws = wb.active

        # 設定標題格式
        header_font = Font(bold=True, color="FFFFFF")  # 白色字體
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # 深藍色背景
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for col_cells in ws.iter_cols(min_row=1, max_row=1):
            for cell in col_cells:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

        # 識別月末日期的欄位索引
        def get_month_end_columns():
            """獲取月末日期對應的欄位索引"""
            month_end_cols = []
            # 從第3欄開始檢查 (跳過Part_No和Model欄位)
            for col in range(3, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value and isinstance(cell_value, datetime):
                    # 檢查是否為月末最後一天
                    date_obj = cell_value.date()
                    last_day_of_month = calendar.monthrange(date_obj.year, date_obj.month)[1]
                    if date_obj.day == last_day_of_month:
                        month_end_cols.append(col)
            return month_end_cols

        month_end_cols = get_month_end_columns()
        print(f"月末欄位索引: {month_end_cols}")

        # 設定內容格式
        light_gray_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")  # 交錯淺灰色背景
        right_align = Alignment(horizontal="right", vertical="center")  # 數字靠右
        left_align = Alignment(horizontal="left", vertical="center")    # 文字靠左
        center_align = Alignment(horizontal="center", vertical="center")  # 居中

        # 負數紅色字體格式
        negative_font = Font(color="FF0000")  # 紅色字體
        negative_bold_font = Font(color="FF0000", bold=True)  # 紅色粗體字體
        
        # 月末粗體格式
        bold_font = Font(bold=True)

        row_count = ws.max_row
        col_count = ws.max_column

        for row in range(2, row_count + 1):  # 跳過標題
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border  # 加邊框

                if row % 2 == 0:  # 偶數行加淺灰色背景
                    cell.fill = light_gray_fill

                if isinstance(cell.value, (int, float)):  # 數值欄位
                    cell.alignment = right_align
                    cell.number_format = "#,##0"
                    
                    # 檢查是否為負數
                    is_negative = cell.value < 0
                    
                    # 檢查是否為月末欄位
                    is_month_end = col in month_end_cols
                    
                    # 根據條件設定字體格式
                    if is_negative and is_month_end:
                        cell.font = negative_bold_font  # 負數且月末：紅色粗體
                    elif is_negative:
                        cell.font = negative_font  # 僅負數：紅色
                    elif is_month_end:
                        cell.font = bold_font  # 僅月末：粗體
                    
                else:
                    cell.alignment = left_align

        # 自動調整欄寬
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

        # 凍結標題列
        ws.freeze_panes = "A2"

        # 儲存格式化後的 Excel
        wb.save(full_path)

        print(f"Daily inventory report exported and formatted: {full_path}")
        print(f"已套用特殊格式:")
        print(f"  - 月末日期欄位: 粗體")
        print(f"  - 負數值: 紅色字體")
        print(f"  - 月末負數值: 紅色粗體")

    except Exception as e:
        logging.error(f"Error exporting to Excel: {e}")
        raise

# Main function
def main()-> int:
    try:
        # Fetch data
        data = fetch_data()
        factory_data = data['factory_data']
        order_data = data['order_data']
        product_data = data['product_data']

        # Calculate daily inventory
        inventory = calculate_inventory(factory_data, order_data, product_data)
        
        # Export to Excel
        export_to_excel(inventory,product_data)
        return 0
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        sys.exit(1)  # Exit the program with a non-zero status code
        return 1

#直接執行的Py的時候
if __name__ == "__main__":
    try:
        # Fetch data
        data = fetch_data()
        factory_data = data['factory_data']
        order_data = data['order_data']
        product_data = data['product_data']

        # Calculate daily inventory
        inventory = calculate_inventory(factory_data, order_data, product_data)
        
        # Export to Excel
        export_to_excel(inventory,product_data)
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        sys.exit(1)  # Exit the program with a non-zero status code